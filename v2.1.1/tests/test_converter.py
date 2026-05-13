"""Tests for ExcelConverter (issue #16, Phase 2).

bytes / 文字列 / Path 入力を受け、ConversionConfig に従って Markdown を
返すことを検証する。runner.run の薄いラッパーとしての振る舞いに重点を置く。
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from excel2md.config import ConversionConfig
from excel2md.converter import ExcelConverter
from excel2md.exceptions import ExcelConversionError, WorkbookOpenError


# =============================================================
# Fixtures
# =============================================================

def _build_small_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "h1"
    ws["B1"] = "h2"
    ws["A2"] = "v1"
    ws["B2"] = "v2"
    wb.save(path)


def _build_xlsx_with_link(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "LinkCell"
    ws["A2"].hyperlink = "https://example.com"
    ws["B2"] = "ValueCell"
    wb.save(path)


# =============================================================
# Construction
# =============================================================

class TestConstruction:
    def test_default_config(self):
        c = ExcelConverter()
        assert isinstance(c.config, ConversionConfig)
        # 既定では CLI 既定と同じ
        assert c.config.csv_markdown_enabled is True
        assert c.config.hyperlink_mode == "footnote"

    def test_custom_config(self):
        cfg = ConversionConfig(hyperlink_mode="inline", max_cells_per_table=10)
        c = ExcelConverter(cfg)
        assert c.config.hyperlink_mode == "inline"
        assert c.config.max_cells_per_table == 10


# =============================================================
# convert() — 通常 Markdown モード
# =============================================================

class TestConvertNormalMarkdown:
    """csv_markdown_enabled=False で通常 Markdown 出力モードの動作を見る。"""

    def test_from_path_str(self, tmp_path):
        xlsx = tmp_path / "in.xlsx"
        _build_small_xlsx(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False)
        result = ExcelConverter(cfg).convert(str(xlsx))
        assert {"markdown", "output_path", "result"}.issubset(result.keys())
        assert result["markdown"]
        assert "h1" in result["markdown"]
        assert "v1" in result["markdown"]

    def test_from_path_obj(self, tmp_path):
        xlsx = tmp_path / "in.xlsx"
        _build_small_xlsx(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False)
        result = ExcelConverter(cfg).convert(xlsx)  # Path object
        assert "h1" in result["markdown"]

    def test_from_bytes(self, tmp_path):
        xlsx = tmp_path / "in.xlsx"
        _build_small_xlsx(xlsx)
        data = xlsx.read_bytes()
        cfg = ConversionConfig(csv_markdown_enabled=False)
        result = ExcelConverter(cfg).convert(data)
        assert "h1" in result["markdown"]
        assert "v1" in result["markdown"]

    def test_bytes_and_path_give_same_markdown(self, tmp_path):
        xlsx = tmp_path / "same.xlsx"
        _build_small_xlsx(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False)
        out_path = ExcelConverter(cfg).convert(str(xlsx))["markdown"]
        out_bytes = ExcelConverter(cfg).convert(xlsx.read_bytes())["markdown"]
        # ファイル名 (input_path.stem) が違うので 1 行目の "変換結果: <name>"
        # は異なる。それ以外は同じはず。
        def _strip_header(md):
            return "\n".join(l for l in md.splitlines() if not l.startswith("# 変換結果"))
        assert _strip_header(out_path) == _strip_header(out_bytes)

    def test_explicit_output_path(self, tmp_path):
        xlsx = tmp_path / "in.xlsx"
        _build_small_xlsx(xlsx)
        out = tmp_path / "explicit.md"
        cfg = ConversionConfig(csv_markdown_enabled=False)
        result = ExcelConverter(cfg).convert(str(xlsx), output_path=str(out))
        assert out.exists()
        assert "h1" in out.read_text(encoding="utf-8")
        assert Path(result["output_path"]) == out


# =============================================================
# convert() — custom config の効果
# =============================================================

class TestCustomConfigEffect:
    def test_hyperlink_mode_inline(self, tmp_path):
        xlsx = tmp_path / "links.xlsx"
        _build_xlsx_with_link(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False, hyperlink_mode="inline")
        md = ExcelConverter(cfg).convert(str(xlsx))["markdown"]
        # inline モードでは URL がセル内に組み込まれる
        # (md_escape で記号類が \ エスケープされるため部分一致で検査)
        assert "https" in md and "example" in md
        # 脚注リファレンスは出ない
        assert "[^" not in md

    def test_hyperlink_mode_footnote(self, tmp_path):
        xlsx = tmp_path / "links.xlsx"
        _build_xlsx_with_link(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False, hyperlink_mode="footnote")
        md = ExcelConverter(cfg).convert(str(xlsx))["markdown"]
        assert "[^1]" in md
        assert "https://example.com" in md

    def test_max_cells_per_table_truncation_does_not_crash(self, tmp_path):
        """Issue #24 の修正と組み合わさり、ExcelConverter 経由でも crash しない。"""
        xlsx = tmp_path / "big.xlsx"
        _build_small_xlsx(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False, max_cells_per_table=2)
        result = ExcelConverter(cfg).convert(str(xlsx))
        assert result["markdown"]  # 何らかの中身が返る


# =============================================================
# convert() — CSV Markdown モード (default)
# =============================================================

class TestConvertCsvMarkdownMode:
    """csv_markdown_enabled=True の経路で markdown が読み返せること。"""

    def test_default_config_returns_markdown(self, tmp_path):
        xlsx = tmp_path / "csv.xlsx"
        _build_small_xlsx(xlsx)
        result = ExcelConverter().convert(str(xlsx))
        # CSV Markdown でも markdown キーは何かしらの中身を持つ
        assert isinstance(result["markdown"], str)
        # 値が空でないことを確認 (path が解決できているということ)
        assert result["markdown"], "expected non-empty markdown content from CSV mode"

    def test_bytes_input_in_csv_mode(self, tmp_path):
        xlsx = tmp_path / "csv.xlsx"
        _build_small_xlsx(xlsx)
        result = ExcelConverter().convert(xlsx.read_bytes())
        assert isinstance(result["markdown"], str)
        assert result["markdown"]


# =============================================================
# Error handling — ライブラリ層は SystemExit ではなく例外を上げる
# (Issue #16 / PR #32 Codex review P1)
# =============================================================

class TestErrorHandling:
    """ライブラリとして呼ばれた場合、xlsx 読み込み失敗で
    プロセスを落とさず ``WorkbookOpenError`` を上げる。"""

    def test_missing_path_raises_workbook_open_error(self, tmp_path):
        missing = tmp_path / "does-not-exist.xlsx"
        with pytest.raises(WorkbookOpenError):
            ExcelConverter().convert(str(missing))

    def test_missing_path_does_not_call_sys_exit(self, tmp_path):
        """SystemExit に化けていないことを明示的に検査する。"""
        missing = tmp_path / "nope.xlsx"
        try:
            ExcelConverter().convert(str(missing))
        except SystemExit:
            pytest.fail("ライブラリ層が SystemExit を投げてはいけない")
        except WorkbookOpenError:
            pass

    def test_corrupt_bytes_raises_workbook_open_error(self):
        with pytest.raises(WorkbookOpenError):
            ExcelConverter().convert(b"this is not a valid xlsx file")

    def test_workbook_open_error_is_excel_conversion_error(self, tmp_path):
        """基底 ``ExcelConversionError`` で一括捕捉できる。"""
        missing = tmp_path / "nope.xlsx"
        with pytest.raises(ExcelConversionError):
            ExcelConverter().convert(str(missing))


class TestCliExitBehavior:
    """CLI は従来どおり exit code 2 で死ぬ (例外はライブラリ層で握り、
    cli.main が exit code に変換する)。"""

    def test_cli_main_exits_with_code_2_on_missing_file(self, tmp_path, capsys):
        from excel2md.cli import main
        missing = tmp_path / "no.xlsx"
        with pytest.raises(SystemExit) as exc_info:
            main([str(missing)])
        assert exc_info.value.code == 2
        captured = capsys.readouterr()
        assert "Failed to open workbook" in captured.err
