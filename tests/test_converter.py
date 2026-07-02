"""Tests for ExcelConverter (issue #16, Phase 2).

bytes / 文字列 / Path 入力を受け、ConversionConfig に従って Markdown を
返すことを検証する。runner.run の薄いラッパーとしての振る舞いに重点を置く。
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from excel2md.config import ConversionConfig
from excel2md.converter import ExcelConverter
from excel2md.exceptions import ExcelConversionError, WorkbookOpenError


# =============================================================
# Fixtures
# =============================================================

def _build_small_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "h1"
    ws["B1"] = "h2"
    ws["A2"] = "v1"
    ws["B2"] = "v2"
    wb.save(path)


def _build_xlsx_with_link(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "LinkCell"
    ws["A2"].hyperlink = "https://example.com"
    ws["B2"] = "ValueCell"
    wb.save(path)


# =============================================================
# Construction
# =============================================================

class TestConstruction:
    def test_default_config(self):
        c = ExcelConverter()
        assert isinstance(c.config, ConversionConfig)
        # 既定では CLI 既定と同じ
        assert c.config.csv_markdown_enabled is True
        assert c.config.hyperlink_mode == "footnote"

    def test_custom_config(self):
        cfg = ConversionConfig(hyperlink_mode="inline", max_cells_per_table=10)
        c = ExcelConverter(cfg)
        assert c.config.hyperlink_mode == "inline"
        assert c.config.max_cells_per_table == 10


# =============================================================
# convert() — 通常 Markdown モード
# =============================================================

class TestConvertNormalMarkdown:
    """csv_markdown_enabled=False で通常 Markdown 出力モードの動作を見る。"""

    def test_from_path_str(self, tmp_path):
        xlsx = tmp_path / "in.xlsx"
        _build_small_xlsx(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False)
        result = ExcelConverter(cfg).convert(str(xlsx))
        assert {"markdown", "output_path", "result"}.issubset(result.keys())
        assert result["markdown"]
        assert "h1" in result["markdown"]
        assert "v1" in result["markdown"]

    def test_from_path_obj(self, tmp_path):
        xlsx = tmp_path / "in.xlsx"
        _build_small_xlsx(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False)
        result = ExcelConverter(cfg).convert(xlsx)  # Path object
        assert "h1" in result["markdown"]

    def test_from_bytes(self, tmp_path):
        xlsx = tmp_path / "in.xlsx"
        _build_small_xlsx(xlsx)
        data = xlsx.read_bytes()
        cfg = ConversionConfig(csv_markdown_enabled=False)
        result = ExcelConverter(cfg).convert(data)
        assert "h1" in result["markdown"]
        assert "v1" in result["markdown"]

    def test_bytes_and_path_give_same_markdown(self, tmp_path):
        xlsx = tmp_path / "same.xlsx"
        _build_small_xlsx(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False)
        out_path = ExcelConverter(cfg).convert(str(xlsx))["markdown"]
        out_bytes = ExcelConverter(cfg).convert(xlsx.read_bytes())["markdown"]
        # ファイル名 (input_path.stem) が違うので 1 行目の "変換結果: <name>"
        # は異なる。それ以外は同じはず。
        def _strip_header(md):
            return "\n".join(l for l in md.splitlines() if not l.startswith("# 変換結果"))
        assert _strip_header(out_path) == _strip_header(out_bytes)

    def test_explicit_output_path(self, tmp_path):
        xlsx = tmp_path / "in.xlsx"
        _build_small_xlsx(xlsx)
        out = tmp_path / "explicit.md"
        cfg = ConversionConfig(csv_markdown_enabled=False)
        result = ExcelConverter(cfg).convert(str(xlsx), output_path=str(out))
        assert out.exists()
        assert "h1" in out.read_text(encoding="utf-8")
        assert Path(result["output_path"]) == out


# =============================================================
# convert() — custom config の効果
# =============================================================

class TestCustomConfigEffect:
    def test_hyperlink_mode_inline(self, tmp_path):
        xlsx = tmp_path / "links.xlsx"
        _build_xlsx_with_link(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False, hyperlink_mode="inline")
        md = ExcelConverter(cfg).convert(str(xlsx))["markdown"]
        # inline モードでは URL がセル内に組み込まれる
        # (md_escape で記号類が \ エスケープされるため部分一致で検査)
        assert "https" in md and "example" in md
        # 脚注リファレンスは出ない
        assert "[^" not in md

    def test_hyperlink_mode_footnote(self, tmp_path):
        xlsx = tmp_path / "links.xlsx"
        _build_xlsx_with_link(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False, hyperlink_mode="footnote")
        md = ExcelConverter(cfg).convert(str(xlsx))["markdown"]
        assert "[^1]" in md
        assert "https://example.com" in md

    def test_max_cells_per_table_truncation_does_not_crash(self, tmp_path):
        """Issue #24 の修正と組み合わさり、ExcelConverter 経由でも crash しない。"""
        xlsx = tmp_path / "big.xlsx"
        _build_small_xlsx(xlsx)
        cfg = ConversionConfig(csv_markdown_enabled=False, max_cells_per_table=2)
        result = ExcelConverter(cfg).convert(str(xlsx))
        assert result["markdown"]  # 何らかの中身が返る


# =============================================================
# convert() — CSV Markdown モード (default)
# =============================================================

class TestConvertCsvMarkdownMode:
    """csv_markdown_enabled=True の経路で markdown が読み返せること。"""

    def test_default_config_returns_markdown(self, tmp_path):
        xlsx = tmp_path / "csv.xlsx"
        _build_small_xlsx(xlsx)
        result = ExcelConverter().convert(str(xlsx))
        # CSV Markdown でも markdown キーは何かしらの中身を持つ
        assert isinstance(result["markdown"], str)
        # 値が空でないことを確認 (path が解決できているということ)
        assert result["markdown"], "expected non-empty markdown content from CSV mode"

    def test_bytes_input_in_csv_mode(self, tmp_path):
        xlsx = tmp_path / "csv.xlsx"
        _build_small_xlsx(xlsx)
        result = ExcelConverter().convert(xlsx.read_bytes())
        assert isinstance(result["markdown"], str)
        assert result["markdown"]


# =============================================================
# Issue #36: ライブラリ層は SystemExit を出さず、例外を上げる
# =============================================================


class TestErrorHandling:
    """ライブラリ経由 (ExcelConverter / convert_to_markdown) で workbook
    オープン失敗が起きた場合、SystemExit ではなく WorkbookOpenError を
    伝播することを保証する。"""

    def test_missing_path_raises_workbook_open_error(self, tmp_path):
        missing = tmp_path / "does-not-exist.xlsx"
        with pytest.raises(WorkbookOpenError):
            ExcelConverter().convert(str(missing))

    def test_corrupt_bytes_raise_workbook_open_error(self):
        with pytest.raises(WorkbookOpenError):
            ExcelConverter().convert(b"not a real xlsx file")

    def test_does_not_raise_system_exit(self, tmp_path):
        """SystemExit に化けないことを明示的に保証する (Issue #36 リグレッション)。"""
        missing = tmp_path / "does-not-exist.xlsx"
        try:
            ExcelConverter().convert(str(missing))
        except SystemExit:
            pytest.fail("ExcelConverter.convert() must not raise SystemExit")
        except WorkbookOpenError:
            pass  # expected

    def test_catchable_by_base_class(self, tmp_path):
        """ExcelConversionError 基底で一括捕捉できる。"""
        missing = tmp_path / "does-not-exist.xlsx"
        with pytest.raises(ExcelConversionError):
            ExcelConverter().convert(str(missing))

    def test_original_exception_is_chained(self, tmp_path):
        """``raise ... from e`` で元の openpyxl 例外が ``__cause__`` に
        保持されている。"""
        missing = tmp_path / "does-not-exist.xlsx"
        with pytest.raises(WorkbookOpenError) as excinfo:
            ExcelConverter().convert(str(missing))
        assert excinfo.value.__cause__ is not None

    def test_convert_to_markdown_propagates_error(self, tmp_path):
        """ワンショット API (convert_to_markdown) でも同じ例外が上がる。"""
        from excel2md import convert_to_markdown

        missing = tmp_path / "does-not-exist.xlsx"
        with pytest.raises(WorkbookOpenError):
            convert_to_markdown(str(missing))


class TestCliExitBehavior:
    """CLI 経路 (cli.main) からは exit code 2 で死ぬという従来挙動を維持する
    (Issue #36 では CLI ユーザー体験を変えない方針)。"""

    def test_cli_exits_with_code_2_on_missing_file(self, tmp_path, capsys):
        from excel2md.cli import main

        missing = tmp_path / "does-not-exist.xlsx"
        with pytest.raises(SystemExit) as excinfo:
            main([str(missing)])
        assert excinfo.value.code == 2
        captured = capsys.readouterr()
        assert "[ERROR]" in captured.err
        assert "Failed to open workbook" in captured.err

    def test_cli_exits_with_code_2_on_corrupt_file(self, tmp_path, capsys):
        from excel2md.cli import main

        corrupt = tmp_path / "broken.xlsx"
        corrupt.write_bytes(b"not a real xlsx file")
        with pytest.raises(SystemExit) as excinfo:
            main([str(corrupt)])
        assert excinfo.value.code == 2
        captured = capsys.readouterr()
        assert "[ERROR]" in captured.err
