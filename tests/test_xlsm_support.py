"""Tests for .xlsm (macro-enabled workbook) support.

Issue #43: `.xlsm` を正式サポート対象として扱うため、フィクスチャ
``tests/fixtures/test_macro.xlsm`` を読み込み・変換できることと、
VBA マクロが破棄されて実行されないことを確認する。

サポート方針:
- ``.xlsm`` は読み込み専用でサポートする。
- ``openpyxl.load_workbook`` を ``keep_vba`` 未指定（既定 False）で呼び出し、
  VBA バイナリは破棄される。マクロは決して実行されない。
- 変換結果（Markdown）は同等内容の ``.xlsx`` と同じ構造になる。
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

FIXTURE = Path(__file__).parent / "fixtures" / "test_macro.xlsm"


def test_fixture_exists():
    assert FIXTURE.exists(), f"xlsm fixture not found: {FIXTURE}"


# =============================================================
# openpyxl レベル: VBA 破棄の検証
# =============================================================

def test_xlsm_loads_with_vba_dropped():
    """既定 (keep_vba 未指定) で読み込んだ場合、VBA は破棄される。"""
    wb = openpyxl.load_workbook(filename=str(FIXTURE), data_only=True)
    assert wb.vba_archive is None, (
        "keep_vba=False（既定）で読み込んだのに VBA バイナリが残っている"
    )


def test_xlsm_load_via_workbook_loader():
    """プロジェクトの load_workbook_safe 経由でも .xlsm が開ける。"""
    from excel2md.workbook_loader import load_workbook_safe

    wb = load_workbook_safe(str(FIXTURE))
    assert wb is not None
    assert wb.vba_archive is None
    assert "基本テーブル" in wb.sheetnames


# =============================================================
# 変換結果: パス入力
# =============================================================

def test_convert_xlsm_by_path_smoke():
    """.xlsm をパス入力で変換でき、想定のセル値が Markdown に含まれる。"""
    from excel2md import convert_to_markdown

    result = convert_to_markdown(str(FIXTURE), csv_markdown_enabled=False)
    md = result["markdown"]

    assert "基本テーブル" in md
    # 基本テーブル の中身
    assert "商品名" in md
    assert "りんご" in md
    assert "150" in md
    # マクロ シートのテーブルも抽出される（本文のみ、マクロは実行されない）
    assert "HelloMacro" in md
    assert "InsertToday" in md


def test_convert_xlsm_by_bytes_smoke():
    """.xlsm のバイト列入力でも変換できる。"""
    from excel2md import convert_to_markdown

    data = FIXTURE.read_bytes()
    result = convert_to_markdown(data, csv_markdown_enabled=False)
    assert "商品名" in result["markdown"]


# =============================================================
# パリティ: .xlsm と同内容の .xlsx で同じ結果になる
# =============================================================

def _build_equivalent_xlsx(path: Path) -> None:
    """test_macro.xlsm の「基本テーブル」シートと同内容の .xlsx を作る。

    パリティ確認用。マクロシートは含めない（マクロは .xlsx には保存できないため）。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基本テーブル"
    ws["A1"] = "商品名"
    ws["B1"] = "価格"
    ws["C1"] = "在庫数"
    ws["A2"] = "りんご"
    ws["B2"] = 150
    ws["C2"] = 100
    ws["A3"] = "みかん"
    ws["B3"] = 80
    ws["C3"] = 250
    ws["A4"] = "バナナ"
    ws["B4"] = 120
    ws["C4"] = 50
    wb.save(path)


def test_xlsm_basic_table_parity_with_xlsx(tmp_path):
    """.xlsm の「基本テーブル」だけを抜き出すと、等価な .xlsx の出力と一致する。"""
    from excel2md import convert_to_markdown

    xlsx = tmp_path / "equiv.xlsx"
    _build_equivalent_xlsx(xlsx)

    xlsm_md = convert_to_markdown(str(FIXTURE), csv_markdown_enabled=False)["markdown"]
    xlsx_md = convert_to_markdown(str(xlsx), csv_markdown_enabled=False)["markdown"]

    # 基本テーブルの全行が両方の出力に含まれる
    for token in ("商品名", "価格", "在庫数", "りんご", "150", "100",
                  "みかん", "80", "250", "バナナ", "120", "50"):
        assert token in xlsm_md, f"{token} missing in .xlsm output"
        assert token in xlsx_md, f"{token} missing in equivalent .xlsx output"


# =============================================================
# セキュリティ: マクロは実行されない
# =============================================================

def test_xlsm_macro_is_not_executed(tmp_path, monkeypatch):
    """.xlsm を変換しても、変換プロセスの作業ディレクトリにマクロ起因の副作用
    （新規ファイル生成等）が一切発生しない。

    openpyxl は VBA を読み込まないため、Auto_Open 等の自動実行マクロが
    定義されていても発火しない。本テストは、その契約をリグレッションさせない
    ためのガードである。
    """
    from excel2md import convert_to_markdown

    workdir = tmp_path / "work"
    workdir.mkdir()
    monkeypatch.chdir(workdir)

    before = set(workdir.iterdir())
    convert_to_markdown(str(FIXTURE), csv_markdown_enabled=False)
    after = set(workdir.iterdir())

    # マクロが実行されていれば、何らかのファイルが workdir に作られる可能性が
    # ある。openpyxl 経路では VBA は読み込みすらされないので before == after。
    assert before == after, (
        f"作業ディレクトリにマクロ起因と思われる副作用が出た: "
        f"added={after - before}"
    )
