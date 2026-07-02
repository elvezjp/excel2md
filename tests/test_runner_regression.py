"""Regression tests for runner-level bugs.

Issue #24: extract_table truncation path must return a 4-tuple (was 3-tuple).
Issue #25: footnote numbering must be unique across tables within the configured scope.
Issue #14: CSV markdown must not expand the print area to cover image positions.
"""
import re
import sys
import tempfile
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))

from excel2md import runner as runner_module
from excel2md.cli import build_argparser
from excel2md.runner import run


def _parse_args(argv):
    return build_argparser().parse_args(argv)


def _make_two_table_workbook(path: Path, with_links: bool = True) -> None:
    """Two tables on one sheet, separated by an empty row.

    When ``with_links`` is True, each table contains exactly one external
    hyperlink so footnote numbering is exercised.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "Link1"
    if with_links:
        ws["A2"].hyperlink = "https://example.com/a"
    ws["B2"] = "Data1"
    # Row 3 is empty -> tables split
    ws["A4"] = "Header3"
    ws["B4"] = "Header4"
    ws["A5"] = "Link2"
    if with_links:
        ws["A5"].hyperlink = "https://example.com/b"
    ws["B5"] = "Data2"
    wb.save(path)


def _make_multi_sheet_workbook(path: Path) -> None:
    """Two sheets, each with a table containing a hyperlink."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "S1"
    ws1["A1"] = "H1"
    ws1["B1"] = "H2"
    ws1["A2"] = "L1"
    ws1["A2"].hyperlink = "https://example.com/s1"
    ws1["B2"] = "V1"

    ws2 = wb.create_sheet("S2")
    ws2["A1"] = "H1"
    ws2["B1"] = "H2"
    ws2["A2"] = "L2"
    ws2["A2"].hyperlink = "https://example.com/s2"
    ws2["B2"] = "V2"
    wb.save(path)


# ============================================================
# Issue #24: truncation return tuple must be 4 elements
# ============================================================

class TestIssue24Truncation:
    """runner.run() must not crash when max_cells_per_table forces truncation."""

    def test_truncation_does_not_crash_runner(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = Path(tmpdir) / "trunc.xlsx"
            _make_two_table_workbook(xlsx, with_links=False)
            out_md = Path(tmpdir) / "out.md"

            args = _parse_args([
                str(xlsx),
                "-o", str(out_md),
                "--no-csv-markdown-enabled",
                "--max-cells-per-table", "2",
            ])

            # Pre-fix this raised ValueError: not enough values to unpack
            result = run(str(xlsx), str(out_md), args)
            assert result is not None
            assert Path(result).exists()
            text = Path(result).read_text(encoding="utf-8")
            assert "max_cells_per_table" in text


# ============================================================
# Issue #25: footnote IDs unique across tables
# ============================================================

_REF_RE = re.compile(r"\[\^(\d+)\]")
_DEF_RE = re.compile(r"^\[\^(\d+)\]:\s*(.+)$", re.MULTILINE)


def _footnote_refs_and_defs(text: str):
    refs = [int(n) for n in _REF_RE.findall(text)]
    defs = [(int(n), body.strip()) for n, body in _DEF_RE.findall(text)]
    return refs, defs


class TestIssue25FootnoteNumbering:
    """Footnote numbers must be unique and sequential within the configured scope."""

    def test_book_scope_footnotes_are_unique_and_sequential(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = Path(tmpdir) / "book.xlsx"
            _make_two_table_workbook(xlsx, with_links=True)
            out_md = Path(tmpdir) / "out.md"

            args = _parse_args([
                str(xlsx),
                "-o", str(out_md),
                "--no-csv-markdown-enabled",
                "--hyperlink-mode", "footnote",
                "--footnote-scope", "book",
            ])
            run(str(xlsx), str(out_md), args)

            text = out_md.read_text(encoding="utf-8")
            refs, defs = _footnote_refs_and_defs(text)

            # Two hyperlinks -> two distinct footnote refs (in body and definitions)
            assert sorted(set(refs)) == [1, 2], f"refs={refs}"
            # Definitions must be unique and contain the two distinct URLs
            assert sorted(n for n, _ in defs) == [1, 2]
            def_urls = {body for _, body in defs}
            assert def_urls == {"https://example.com/a", "https://example.com/b"}

    def test_sheet_scope_resets_per_sheet(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = Path(tmpdir) / "sheet.xlsx"
            _make_multi_sheet_workbook(xlsx)
            out_md = Path(tmpdir) / "out.md"

            args = _parse_args([
                str(xlsx),
                "-o", str(out_md),
                "--no-csv-markdown-enabled",
                "--hyperlink-mode", "footnote",
                "--footnote-scope", "sheet",
            ])
            run(str(xlsx), str(out_md), args)

            text = out_md.read_text(encoding="utf-8")
            refs, defs = _footnote_refs_and_defs(text)

            # Each sheet has one link; with sheet scope each sheet starts at 1.
            assert refs.count(1) >= 2, f"refs={refs}"
            assert defs == [
                (1, "https://example.com/s1"),
                (1, "https://example.com/s2"),
            ]


# ============================================================
# Issue #14: CSV markdown must respect the print area, not expand
# the union_area to cover image positions outside it.
# ============================================================


def _make_workbook_with_print_area(path: Path) -> None:
    """Workbook whose print area is A1:B2 but with data also at C5/D6.

    Used to verify the CSV markdown range stays inside the print area even
    when images are reported at out-of-area positions.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "h1"
    ws["B1"] = "h2"
    ws["A2"] = "v1"
    ws["B2"] = "v2"
    # Cells outside the print area
    ws["C5"] = "leak-c5"
    ws["D6"] = "leak-d6"
    ws.print_area = "A1:B2"
    wb.save(path)


class TestIssue14CsvPrintAreaRespect:
    """CSV markdown output must stay within the declared print area."""

    def test_out_of_area_image_does_not_expand_range(self, monkeypatch):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = Path(tmpdir) / "area.xlsx"
            _make_workbook_with_print_area(xlsx)

            inside_path = "out/inside.png"
            outside_path = "out/outside.png"

            def fake_extract_images(ws, output_dir, sheet_name, md_basename, opts, xlsx_path=None):
                # One image inside the print area (B2) and one outside (E5).
                return {(2, 2): inside_path, (5, 5): outside_path}

            monkeypatch.setattr(runner_module, "extract_images_from_sheet", fake_extract_images)

            args = _parse_args([
                str(xlsx),
                "--csv-output-dir", tmpdir,
            ])
            result = run(str(xlsx), None, args)
            assert result is not None

            out_files = list(Path(tmpdir).glob("*.md"))
            assert out_files, "expected at least one CSV markdown output file"
            text = "\n".join(p.read_text(encoding="utf-8") for p in out_files)

            # The declared range must remain the print area.
            assert "A1:B2" in text, text
            # The inside image must appear; the outside one must NOT.
            assert inside_path in text
            assert outside_path not in text
            # Out-of-area data cells must not leak into the CSV output either.
            assert "leak-c5" not in text
            assert "leak-d6" not in text


class TestIssue26CsvOnlySkipsNormalMarkdown:
    """Issue #26: the default CSV-only path must not run the normal-Markdown
    pipeline (table detection/extraction/formatting), and shapes-mode Mermaid
    extraction must run at most once per sheet."""

    def test_csv_default_skips_table_detection_and_extraction(self, monkeypatch):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = Path(tmpdir) / "book.xlsx"
            _make_two_table_workbook(xlsx, with_links=False)

            def fail_grid_to_tables(*a, **kw):
                raise AssertionError("grid_to_tables must not run in CSV-only mode")

            def fail_extract_table(*a, **kw):
                raise AssertionError("extract_table must not run in CSV-only mode")

            monkeypatch.setattr(runner_module, "grid_to_tables", fail_grid_to_tables)
            monkeypatch.setattr(runner_module, "extract_table", fail_extract_table)

            args = _parse_args([str(xlsx), "--csv-output-dir", tmpdir])
            result = run(str(xlsx), None, args)
            assert result is not None

            out_files = list(Path(tmpdir).glob("*_csv.md"))
            assert out_files, "expected CSV markdown output file"
            text = out_files[0].read_text(encoding="utf-8")
            assert "Data1" in text
            assert "Data2" in text

    def test_normal_markdown_mode_still_runs_pipeline(self, monkeypatch):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = Path(tmpdir) / "book.xlsx"
            _make_two_table_workbook(xlsx, with_links=False)

            calls = {"grid_to_tables": 0}
            real_grid_to_tables = runner_module.grid_to_tables

            def counting_grid_to_tables(*a, **kw):
                calls["grid_to_tables"] += 1
                return real_grid_to_tables(*a, **kw)

            monkeypatch.setattr(runner_module, "grid_to_tables", counting_grid_to_tables)

            out_md = Path(tmpdir) / "out.md"
            args = _parse_args([str(xlsx), "--no-csv-markdown-enabled"])
            result = run(str(xlsx), str(out_md), args)
            assert result == str(out_md)
            assert calls["grid_to_tables"] > 0
            assert "Data1" in out_md.read_text(encoding="utf-8")

    def test_shapes_mermaid_extracted_once_per_sheet(self, monkeypatch):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = Path(tmpdir) / "book.xlsx"
            _make_two_table_workbook(xlsx, with_links=False)

            calls = {"count": 0}

            def fake_shapes_mermaid(input_path, ws, opts):
                calls["count"] += 1
                return "```mermaid\nflowchart TD\n  A --> B\n```"

            monkeypatch.setattr(runner_module, "_v14_extract_shapes_to_mermaid", fake_shapes_mermaid)

            args = _parse_args([
                str(xlsx),
                "--mermaid-enabled",
                "--csv-output-dir", tmpdir,
            ])
            result = run(str(xlsx), None, args)
            assert result is not None

            # Single sheet: extraction must run exactly once (previously twice —
            # once for the discarded normal Markdown and once for CSV markdown).
            assert calls["count"] == 1

            out_files = list(Path(tmpdir).glob("*_csv.md"))
            assert out_files
            text = out_files[0].read_text(encoding="utf-8")
            assert "```mermaid" in text
