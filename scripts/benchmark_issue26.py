# -*- coding: utf-8 -*-
"""Issue #26 性能改善の効果計測用ベンチマーク。

生成 fixture に対して変換全体 (runner.run) の所要時間を計測する。
改善前後で同一条件の数値を比較できるよう、ワークブック生成は決定的に行う。

使い方:
    uv run python scripts/benchmark_issue26.py            # 全ケース実行
    uv run python scripts/benchmark_issue26.py --repeat 5 # 繰り返し回数変更
    uv run python scripts/benchmark_issue26.py --cases 50k_cells merged_cells

計測ケース:
    50k_cells          500行 × 100列 (Issue #26 記載の「実用上大きめの表」)
    200k_cells         2000行 × 100列 (max_cells_per_table 近辺)
    merged_cells       500行 × 100列 + 結合セル多数 (merged_lookup 構築の負荷確認)
    multi_sheet_mermaid 10シート × 100行 × 100列, mermaid shapes モード有効
                        (シートごとの DrawingML / ZIP 解析の負荷確認)
    mermaid_fixture    tests/fixtures/test_mermaid.xlsx × 30回, mermaid 有効
                        (実図形入りファイルでの DrawingML 解析の負荷確認)
"""

import argparse
import shutil
import statistics
import sys
import tempfile
import time
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from openpyxl import Workbook

from excel2md.config import ConversionConfig
from excel2md.runner import run


def _fill_sheet(ws, rows: int, cols: int, empty_row_interval: int = 50):
    """決定的な内容でシートを埋める。empty_row_interval ごとに空行を入れ、
    通常 Markdown 経路のテーブル分割検出が実際に働くようにする。"""
    for c in range(1, cols + 1):
        ws.cell(row=1, column=c, value=f"Header{c}")
    for r in range(2, rows + 1):
        if empty_row_interval and r % empty_row_interval == 0:
            continue  # 空行 (テーブル分割の境界になる)
        for c in range(1, cols + 1):
            if c % 3 == 0:
                ws.cell(row=r, column=c, value=(r * 31 + c * 7) % 10000)
            elif c % 3 == 1:
                ws.cell(row=r, column=c, value=f"val_{r}_{c}")
            else:
                ws.cell(row=r, column=c, value=f"{(r + c) % 97}.{c % 10}")


def gen_table_workbook(path: Path, rows: int, cols: int, sheets: int = 1,
                       merged: bool = False):
    wb = Workbook()
    for i in range(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = f"Sheet{i + 1}"
        _fill_sheet(ws, rows, cols)
        if merged:
            # 10行ごとに 3列分の結合セルを敷き詰める
            for r in range(2, rows + 1, 10):
                for c in range(1, cols - 3, 6):
                    ws.merge_cells(start_row=r, start_column=c,
                                   end_row=r + 1, end_column=c + 2)
    wb.save(path)


def build_cases(work_dir: Path):
    """ケース名 -> (入力パス, ConversionConfig, 実行回数/計測1回あたり) を返す。"""
    cases = {}

    p = work_dir / "bench_50k.xlsx"
    gen_table_workbook(p, rows=500, cols=100)
    cases["50k_cells"] = (p, ConversionConfig(), 1)

    p = work_dir / "bench_200k.xlsx"
    gen_table_workbook(p, rows=2000, cols=100)
    cases["200k_cells"] = (p, ConversionConfig(), 1)

    p = work_dir / "bench_merged.xlsx"
    gen_table_workbook(p, rows=500, cols=100, merged=True)
    cases["merged_cells"] = (p, ConversionConfig(), 1)

    p = work_dir / "bench_multisheet.xlsx"
    gen_table_workbook(p, rows=100, cols=100, sheets=10)
    cases["multi_sheet_mermaid"] = (p, ConversionConfig(mermaid_enabled=True), 1)

    fixture = REPO_ROOT / "tests" / "fixtures" / "test_mermaid.xlsx"
    if fixture.exists():
        cases["mermaid_fixture"] = (fixture, ConversionConfig(mermaid_enabled=True), 30)

    return cases


def run_case(input_path: Path, config: ConversionConfig, inner_loops: int,
             out_dir: Path) -> float:
    """変換1計測分 (inner_loops 回) の所要秒数を返す。出力は out_dir に捨てる。"""
    config.csv_output_dir = str(out_dir)
    start = time.perf_counter()
    for _ in range(inner_loops):
        run(str(input_path), str(out_dir / "out.md"), config)
    return time.perf_counter() - start


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--repeat", type=int, default=3, help="計測の繰り返し回数 (default: 3)")
    ap.add_argument("--cases", nargs="*", default=None, help="実行するケース名 (default: 全ケース)")
    args = ap.parse_args()

    work_dir = Path(tempfile.mkdtemp(prefix="excel2md_bench_"))
    try:
        print("generating fixtures ...", flush=True)
        cases = build_cases(work_dir)
        if args.cases:
            unknown = set(args.cases) - set(cases)
            if unknown:
                ap.error(f"unknown cases: {sorted(unknown)} (available: {sorted(cases)})")
            cases = {k: v for k, v in cases.items() if k in args.cases}

        print(f"{'case':<22}{'best':>10}{'median':>10}  (repeat={args.repeat})")
        for name, (input_path, config, inner_loops) in cases.items():
            times = []
            for _ in range(args.repeat):
                out_dir = work_dir / f"out_{name}"
                if out_dir.exists():
                    shutil.rmtree(out_dir)
                out_dir.mkdir()
                times.append(run_case(input_path, config, inner_loops, out_dir))
            print(f"{name:<22}{min(times):>9.3f}s{statistics.median(times):>9.3f}s")
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


if __name__ == "__main__":
    main()
