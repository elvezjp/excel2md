"""Regression tests for the public API surface.

Issue #15: ``is_code_block`` and ``build_code_block_from_rows`` were importable
from the top-level ``excel_to_md`` module in v1.8 but were dropped from the
v2.x public surface. They must remain importable from both ``excel2md`` and
``excel_to_md`` for backward compatibility.

Issue #16 (Phase 3): ``ConversionConfig`` / ``ExcelConverter`` /
``convert_to_markdown`` should be exposed as the canonical library API from
both ``excel2md`` and ``excel_to_md``.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))


_JAVA_ROWS = [
    ["public class Example {"],
    ["    private int value;"],
    ["}"],
]


def test_is_code_block_importable_from_excel2md_package():
    from excel2md import is_code_block

    assert callable(is_code_block)
    assert is_code_block(_JAVA_ROWS) is True
    assert is_code_block([["Name", "Value"], ["Item", "1"]]) is False


def test_build_code_block_from_rows_importable_from_excel2md_package():
    from excel2md import build_code_block_from_rows

    assert callable(build_code_block_from_rows)
    out = build_code_block_from_rows(_JAVA_ROWS)
    assert out is not None
    assert "public class Example" in out


def test_is_code_block_importable_from_excel_to_md_facade():
    from excel_to_md import is_code_block

    assert callable(is_code_block)


def test_build_code_block_from_rows_importable_from_excel_to_md_facade():
    from excel_to_md import build_code_block_from_rows

    assert callable(build_code_block_from_rows)


def test_same_callable_across_paths():
    """Re-exports must point at the canonical implementation, not copies."""
    from excel2md import is_code_block as via_pkg
    from excel_to_md import is_code_block as via_facade
    from excel2md.table_formatting import is_code_block as canonical

    assert via_pkg is canonical
    assert via_facade is canonical


# =============================================================
# Issue #16 — library API surface
# =============================================================

def test_conversion_config_importable_from_excel2md():
    from excel2md import ConversionConfig

    cfg = ConversionConfig()
    assert cfg.hyperlink_mode == "footnote"  # default


def test_conversion_config_importable_from_excel_to_md_facade():
    from excel_to_md import ConversionConfig

    cfg = ConversionConfig(max_cells_per_table=10)
    assert cfg.max_cells_per_table == 10


def test_excel_converter_importable_from_excel2md():
    from excel2md import ExcelConverter

    conv = ExcelConverter()
    # 既定値で動作する
    assert conv.config.csv_markdown_enabled is True


def test_excel_converter_importable_from_excel_to_md_facade():
    from excel_to_md import ExcelConverter

    conv = ExcelConverter()
    assert conv is not None


def test_convert_to_markdown_importable_from_excel2md():
    from excel2md import convert_to_markdown

    assert callable(convert_to_markdown)


def test_convert_to_markdown_importable_from_excel_to_md_facade():
    from excel_to_md import convert_to_markdown

    assert callable(convert_to_markdown)


def test_library_api_objects_are_same_across_paths():
    from excel2md import (
        ConversionConfig as CC1,
        ExcelConverter as EC1,
        convert_to_markdown as ctm1,
    )
    from excel_to_md import (
        ConversionConfig as CC2,
        ExcelConverter as EC2,
        convert_to_markdown as ctm2,
    )

    assert CC1 is CC2
    assert EC1 is EC2
    assert ctm1 is ctm2


# =============================================================
# convert_to_markdown — 動作確認 (薄いラッパーであること)
# =============================================================

def _build_tiny_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "h"
    ws["A2"] = "v"
    wb.save(path)


def test_convert_to_markdown_with_path(tmp_path):
    from excel2md import convert_to_markdown

    xlsx = tmp_path / "tiny.xlsx"
    _build_tiny_xlsx(xlsx)
    result = convert_to_markdown(str(xlsx), csv_markdown_enabled=False)
    assert "markdown" in result and result["markdown"]
    assert "h" in result["markdown"]


def test_convert_to_markdown_with_bytes(tmp_path):
    from excel2md import convert_to_markdown

    xlsx = tmp_path / "tiny.xlsx"
    _build_tiny_xlsx(xlsx)
    result = convert_to_markdown(xlsx.read_bytes(), csv_markdown_enabled=False)
    assert "h" in result["markdown"]


def test_convert_to_markdown_kwargs_passed_to_config(tmp_path):
    """config_kwargs が ConversionConfig に正しく渡ること。"""
    from excel2md import convert_to_markdown

    xlsx = tmp_path / "tiny.xlsx"
    _build_tiny_xlsx(xlsx)
    # hyperlink_mode をデフォルトの footnote から inline に変えて違いが出ること
    # (このテスト用 xlsx にはリンクが無いので markdown 内容は変わらないが、
    #  kwargs が引き渡されることを ConversionConfig の TypeError で間接的に
    #  検査するために、わざと未知のキー指定で例外を確認)
    with pytest.raises(TypeError):
        convert_to_markdown(str(xlsx), nonexistent_option=True)


def test_convert_to_markdown_default_csv_mode_returns_string(tmp_path):
    """既定設定 (csv_markdown_enabled=True) でも markdown は文字列で返る。"""
    from excel2md import convert_to_markdown

    xlsx = tmp_path / "tiny.xlsx"
    _build_tiny_xlsx(xlsx)
    result = convert_to_markdown(str(xlsx))
    assert isinstance(result["markdown"], str)
    assert result["markdown"]  # 非空
